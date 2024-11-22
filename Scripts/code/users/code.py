import json
import os.path
import time
import traceback
import openpyxl
import requests
from bs4 import BeautifulSoup

start_time=time.time()

file_path = "D:/"
if os.path.exists(file_path):
    workbook = openpyxl.load_workbook(file_path)
else:
    workbook = openpyxl.Workbook()
sheet = workbook.active


# 请求函数
def request_data(url):
    # 构造一个请求头，请求头根据浏览器变化，没有固定值
    headers = {
        "Accept-Language": "zh-CN,zh;q=0.9",
        "Cache-Control": "max-age=0",
        # "Cookie": "PHPSESSID=mmma2fotv1cujd7cvep5mos096; wp_layer_content_layer8BD4ADB0EEC7F8E8ADF816E21794D7CE=%5B%5D; Hm_lvt_140ecc4b781293a2d289fb7807a8d182=1704551181; __bid_n=18cdf2af74d75361431832; wp_layer_page_layer8BD4ADB0EEC7F8E8ADF816E21794D7CE=1; Hm_lpvt_140ecc4b781293a2d289fb7807a8d182=1704551670",
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/118.0.0.0 Safari/537.36"
    }
    # 尝试语句
    response = requests.get(url=url, headers=headers,
                            timeout=60, proxies={
            'http': None,
            'https': None,
        })

    items = json.loads(response.text)['directory_items']
    for item in items:
        try:
            user = item['user']

            ret = requests.get(url=f"https://community.openai.com/u/{user['username']}.json",
                               headers=headers,
                               timeout=60, proxies={
                    'http': None,
                    'https': None,
                })
            mainUser = ret.json()['user']
            if 'badges' in ret.json():
                badges = ret.json()['badges']
                slug = badges[0]['slug']
            else:
                badges = []
                slug = ''
            ret = requests.get(url=f"https://community.openai.com/u/{user['username']}/summary.json",
                               headers=headers,
                               timeout=60, proxies={
                    'http': None,
                    'https': None,
                })
            user_summary = ret.json()['user_summary']
            max_row = sheet.max_row
            print(mainUser)
            sheet.cell(max_row + 1, 1).value = user['username']
            sheet.cell(max_row + 1, 2).value = user['name']
            sheet.cell(max_row + 1, 3).value = mainUser['created_at']
            sheet.cell(max_row + 1, 4).value = mainUser['last_posted_at']
            sheet.cell(max_row + 1, 5).value = mainUser['last_seen_at']
            sheet.cell(max_row + 1, 6).value = mainUser['profile_view_count']
            sheet.cell(max_row + 1, 7).value = slug
            sheet.cell(max_row + 1, 8).value = user_summary['days_visited']  # days visited
            sheet.cell(max_row + 1, 9).value = user_summary['time_read']  # read time
            sheet.cell(max_row + 1, 10).value = user_summary['recent_time_read']  # recent read time
            sheet.cell(max_row + 1, 11).value = user_summary['topics_entered']  # topics viewed
            sheet.cell(max_row + 1, 12).value = user_summary['posts_read_count']  # posts_read_count
            sheet.cell(max_row + 1, 13).value = user_summary['likes_given']  # given
            sheet.cell(max_row + 1, 14).value = user_summary['likes_received']  # likes_received
            sheet.cell(max_row + 1, 15).value = user_summary['topic_count']  # topic_count
            sheet.cell(max_row + 1, 16).value = user_summary['post_count']  # post_count
            sheet.cell(max_row + 1, 17).value = user_summary['solved_count']  # solved_count
        except Exception as e:
            traceback.print_exc()
    workbook.save(file_path)


if __name__ == '__main__':
    if os.path.exists("D:/张广倍大学/张广倍科研/爬虫新_2024.7/代码汇总/爬用户/page.txt"):
        last_page = open("D:/张广倍大学/张广倍科研/爬虫新_2024.7/代码汇总/爬用户/page.txt", 'r').read()
        page = int(last_page)
    else:
        page = 0
    sheet.cell(1, 1).value = 'username'
    sheet.cell(1, 2).value = 'name'
    sheet.cell(1, 3).value = 'Joined'
    sheet.cell(1, 4).value = 'Last Post'
    sheet.cell(1, 5).value = 'Seen'
    sheet.cell(1, 6).value = 'Views'
    sheet.cell(1, 7).value = 'Trust Level'
    sheet.cell(1, 8).value = 'days visited'
    sheet.cell(1, 9).value = 'read time'
    sheet.cell(1, 10).value = 'recent read time'
    sheet.cell(1, 11).value = 'topics viewed'
    sheet.cell(1, 12).value = 'posts read'
    sheet.cell(1, 13).value = 'given'
    sheet.cell(1, 14).value = 'received'
    sheet.cell(1, 15).value = 'topics created'
    sheet.cell(1, 16).value = 'posts created'
    sheet.cell(1, 17).value = 'solutions'
    while True:
        print(f"第{page}页")
        # 调用构造链接函数，并且接受返回值
        request_data(
            f"https://community.openai.com/directory_items.json?order=likes_received&page={page}&period=all")
        page += 1
        with open("D:/", 'w') as f:
            f.write(f'{page}')

end_time=time.time()
print("over!"+str(end_time-start_time))